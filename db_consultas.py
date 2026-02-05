# db_consultas.py
import pyodbc
import pandas as pd
from typing import List, Dict, Optional, Tuple
import logging
import warnings
from datetime import datetime
import os

from cadenas_config import obtener_cdn_id, validar_cadena, obtener_categorias_excluidas
from credenciales_manager import obtener_connection_string

warnings.filterwarnings('ignore')
logger = logging.getLogger(__name__)

class ConsultasDB:
    def __init__(self):
        self.connection_string = obtener_connection_string()
        self.conexion = None
        if self.connection_string is None:
            logger.error("No se pudo obtener el connection string")
    
    def conectar(self) -> bool:
        if self.connection_string is None:
            logger.error("Connection string no disponible")
            return False
        
        try:
            logger.info("Conectando a la base de datos...")
            self.conexion = pyodbc.connect(self.connection_string)
            logger.info("Conexión exitosa a la base de datos")
            return True
        except Exception as e:
            logger.error(f"Error al conectar a la base de datos: {e}")
            return False
    
    def desconectar(self):
        if self.conexion:
            try:
                self.conexion.close()
                logger.info("Conexión cerrada")
            except Exception as e:
                logger.warning(f"Error al cerrar conexión: {e}")
    
    def obtener_categorias_por_cadena(self, cdn_id: int) -> Optional[pd.DataFrame]:
        try:
            # Obtener categorías excluidas para esta cadena
            categorias_excluidas = obtener_categorias_excluidas(cdn_id)
            
            query = """
            SELECT IDCategoria, cat_abreviatura, cat_descripcion
            FROM Categoria
            WHERE cdn_id = ?
            AND IDStatus = '71039503-85CF-E511-80C6-000D3A3261F3'
            ORDER BY cat_descripcion
            """
            
            logger.info(f"Ejecutando consulta de categorías para cdn_id: {cdn_id}")
            df = pd.read_sql(query, self.conexion, params=(cdn_id,))
            
            if df.empty:
                logger.warning(f"No se encontraron categorías para cdn_id: {cdn_id}")
                return None
            
            # Filtrar categorías excluidas
            if categorias_excluidas:
                df_filtrado = df[~df['IDCategoria'].isin(categorias_excluidas)]
                logger.info(f"Filtradas {len(df) - len(df_filtrado)} categorías excluidas")
                df = df_filtrado
            
            logger.info(f"Se encontraron {len(df)} categorías (después de filtrar excluidas)")
            return df
            
        except Exception as e:
            logger.error(f"Error al obtener categorías: {e}")
            return None
    
    def identificar_columnas_sp(self, df: pd.DataFrame) -> Dict[str, str]:
        """Identifica automáticamente las columnas del stored procedure"""
        columnas_identificadas = {}
        
        for col in df.columns:
            col_str = str(col).lower()
            
            # Identificar PLU_ID
            if 'plu_id' in col_str and 'categoria' not in col_str and 'cat' not in col_str:
                columnas_identificadas['plu_id'] = col
            
            # Identificar PLU_NUM
            elif 'plu_num' in col_str and 'plu' in col_str:
                columnas_identificadas['plu_num'] = col
            
            # Identificar descripción
            elif ('descripcion' in col_str or 'descripción' in col_str) and 'plu' in col_str:
                columnas_identificadas['descripcion'] = col
            
            # Identificar PVP
            elif ('pvp' in col_str or 'pr_pvp' in col_str or 'precio' in col_str):
                columnas_identificadas['pvp'] = col
            
            # Identificar categoría
            elif ('categoria' in col_str or 'cat_id' in col_str or 'idcategoria' in col_str):
                columnas_identificadas['categoria_id'] = col
        
        logger.info(f"Columnas identificadas: {columnas_identificadas}")
        return columnas_identificadas
    
    def ejecutar_stored_procedure_precios(
        self,
        cdn_id: int,
        categorias_df: pd.DataFrame,
        canal_param: str = 'Canal',
        canal_ids: Optional[List[str]] = None
    ) -> Optional[pd.DataFrame]:
        cursor = None
        try:
            # Convertir categorías a lista de IDs
            categorias = categorias_df['IDCategoria'].tolist()
            categorias_str = ','.join(categorias)

            
            # Configurar canal_ids por defecto si no se proporcionan
            if canal_ids is None:
                canal_ids_str = (
    "A457436C-84DE-E711-80D0-000D3A019254,"
    "9B4010B1-C839-E811-80D1-000D3A019254,"
    "0D049503-85CF-E511-80C6-000D3A3261F3,"
    "0E049503-85CF-E511-80C6-000D3A3261F3,"
    "0F049503-85CF-E511-80C6-000D3A3261F3,"
    "10049503-85CF-E511-80C6-000D3A3261F3"
)
            else:
                canal_ids_str = ','.join([f"'{canal}'" for canal in canal_ids])
            
            # Construir query del stored procedure
            query = f"""
            EXEC [config].[USP_administracionPrecios_listadoDePrecios] 
                {cdn_id}, 
                '{categorias_str}', 
                '{canal_param}', 
                '{canal_ids_str}'
            """
            
            logger.info(f"Ejecutando stored procedure para cdn_id: {cdn_id}")
            logger.info(f"Query: {query[:200]}...")  # Log parcial para debugging
            
            cursor = self.conexion.cursor()
            cursor.execute(query)
            
            # Manejar múltiples result sets si es necesario
            try:
                while cursor.description is None:
                    cursor.nextset()
            except pyodbc.ProgrammingError:
                pass
            
            if cursor.description is None:
                logger.warning("El stored procedure no devolvió resultados")
                cursor.close()
                return None
            
            # Obtener datos
            columnas = [col[0] for col in cursor.description]
            datos = cursor.fetchall()
            cursor.close()
            
            df = pd.DataFrame.from_records(datos, columns=columnas)
            
            if df.empty:
                logger.warning("El stored procedure no devolvió datos")
                return None
            
            logger.info(f"Stored procedure ejecutado exitosamente: {len(df)} filas, {len(df.columns)} columnas")
            logger.info(f"Primeras 5 filas:\n{df.head()}")
            
            return df
            
        except pyodbc.Error as e:
            logger.error(f"Error SQL al ejecutar stored procedure:")
            logger.error(f"Código de error: {e.args[0] if e.args else 'N/A'}")
            logger.error(f"Mensaje: {e.args[1] if len(e.args) > 1 else str(e)}")
            if cursor:
                cursor.close()
            return None
            
        except Exception as e:
            logger.error(f"Error general al ejecutar stored procedure: {type(e).__name__}")
            logger.error(f"Detalles: {str(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            if cursor:
                cursor.close()
            return None
    
    def generar_archivo_excel(
        self,
        df: pd.DataFrame,
        categorias_df: pd.DataFrame,
        nombre_cadena: str,
        cdn_id: int = None,
        ruta_salida: str = None
    ) -> Optional[str]:
        try:
            import os
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            import numpy as np
            
            # Si no se proporciona ruta, usar directorio actual del script
            if ruta_salida is None:
                script_dir = os.path.dirname(os.path.abspath(__file__))
                ruta_salida = os.path.join(script_dir, "descargas")
            
            # Convertir a ruta absoluta para evitar problemas
            ruta_salida = os.path.abspath(ruta_salida)
            
            # Crear directorio de descargas si no existe
            os.makedirs(ruta_salida, exist_ok=True)
            
            # Generar nombre de archivo con timestamp único para evitar conflictos
            fecha = datetime.now()
            timestamp = fecha.strftime("%Y%m%d_%H%M%S")
            nombre_archivo = f"Precios_Plu_{nombre_cadena.replace(' ', '_')}_{timestamp}.xlsx"
            ruta_completa = os.path.join(ruta_salida, nombre_archivo)
            
            logger.info(f"Generando archivo Excel: {nombre_archivo}")
            logger.info(f"Datos recibidos del SP: {len(df)} filas, {len(df.columns)} columnas")
            logger.info(f"Columnas disponibles: {list(df.columns)}")
            
            # Identificar columnas del SP
            columnas_sp = self.identificar_columnas_sp(df)
            
            # Verificar que tenemos las columnas necesarias
            columnas_necesarias = ['plu_id', 'descripcion', 'pvp', 'categoria_id']
            columnas_faltantes = [col for col in columnas_necesarias if col not in columnas_sp]
            
            if columnas_faltantes:
                logger.error(f"Faltan columnas necesarias: {columnas_faltantes}")
                logger.error(f"Columnas disponibles: {list(df.columns)}")
                return None
            
            plu_id_col = columnas_sp['plu_id']
            descripcion_col = columnas_sp['descripcion']
            pvp_col = columnas_sp['pvp']
            categoria_id_col = columnas_sp['categoria_id']
            plu_num_col = columnas_sp.get('plu_num', plu_id_col)  # Usar PLU_ID si no hay PLU_NUM
            
            logger.info(f"Columnas mapeadas:")
            logger.info(f"  PLU_NUM: {plu_num_col}")
            logger.info(f"  PLU_ID: {plu_id_col}")
            logger.info(f"  DESCRIPCION: {descripcion_col}")
            logger.info(f"  PVP: {pvp_col}")
            logger.info(f"  CATEGORIA_ID: {categoria_id_col}")
            
            # Limpiar datos
            df_limpio = df.copy()
            
            # Convertir todas las columnas a sus tipos correctos
            df_limpio[plu_id_col] = df_limpio[plu_id_col].astype(str).str.strip().str.upper()
            df_limpio[categoria_id_col] = df_limpio[categoria_id_col].astype(str).str.strip().str.upper()
            df_limpio[pvp_col] = pd.to_numeric(df_limpio[pvp_col], errors='coerce').fillna(0.0)
            
            # Filtrar datos de categorías excluidas
            if cdn_id:
                categorias_excluidas = obtener_categorias_excluidas(cdn_id)
                if categorias_excluidas:
                    registros_originales = len(df_limpio)
                    df_limpio = df_limpio[~df_limpio[categoria_id_col].isin(categorias_excluidas)]
                    logger.info(f"Filtrados {registros_originales - len(df_limpio)} registros de categorías excluidas")
            
            # Log de muestra de datos
            logger.info(f"Muestra de datos limpios (primeras 3 filas):")
            logger.info(f"{df_limpio[[plu_num_col, plu_id_col, categoria_id_col, pvp_col]].head(3)}")
            
            # Obtener lista única de productos (sin categoría)
            columnas_producto = [plu_num_col, plu_id_col, descripcion_col]
            productos_unicos = df_limpio[columnas_producto].drop_duplicates(subset=[plu_id_col]).sort_values(plu_num_col).reset_index(drop=True)
            
            logger.info(f"Total de productos únicos: {len(productos_unicos)}")
            
            # Crear DataFrame final con columnas básicas
            df_final = pd.DataFrame()
            df_final['#PLU_NUM_PLU'] = productos_unicos[plu_num_col]
            df_final['#PLU'] = productos_unicos[plu_id_col]
            df_final['PRODUCTO'] = productos_unicos[descripcion_col]
            
            # Para cada categoría, crear columna con precios
            logger.info(f"Procesando {len(categorias_df)} categorías...")
            
            for idx_cat, cat_row in categorias_df.iterrows():
                cat_id = str(cat_row['IDCategoria']).strip().upper()
                cat_nombre = cat_row['cat_descripcion']
                nombre_columna = f"{cat_nombre}|{cat_id}"
                
                logger.info(f"Procesando categoría: {cat_nombre} (ID: {cat_id})")
                
                # Filtrar datos SOLO de esta categoría específica
                mascara_categoria = df_limpio[categoria_id_col] == cat_id
                datos_categoria = df_limpio[mascara_categoria].copy()
                
                logger.info(f"  Registros encontrados para esta categoría: {len(datos_categoria)}")
                
                # Crear columna de precios para esta categoría
                columna_precios = []
                
                for plu_final in df_final['#PLU']:
                    plu_key = str(plu_final).strip().upper()
                    
                    # Buscar este PLU específico en los datos de esta categoría específica
                    mascara_plu = datos_categoria[plu_id_col] == plu_key
                    filas_plu_en_categoria = datos_categoria[mascara_plu]
                    
                    if len(filas_plu_en_categoria) > 0:
                        # Este PLU existe en esta categoría
                        # Tomar el primer precio encontrado
                        precio_value = filas_plu_en_categoria.iloc[0][pvp_col]
                        
                        # Si es nulo, NaN o None, usar 0.0
                        if pd.isna(precio_value) or precio_value is None:
                            precio = 0.0
                        else:
                            try:
                                precio = float(precio_value)
                                precio = max(0.0, precio)  # No negativos
                            except (ValueError, TypeError):
                                precio = 0.0
                    else:
                        # Este PLU NO existe en esta categoría
                        precio = 0.0
                    
                    columna_precios.append(precio)
                
                df_final[nombre_columna] = columna_precios
                
                # Log de verificación
                precios_no_cero = sum(1 for p in columna_precios if p > 0)
                logger.info(f"  Productos con precio > 0 en columna final: {precios_no_cero}")
            
            # Asegurar formato numérico en todas las columnas de precios
            for col in df_final.columns:
                if '|' in col:  # Es una columna de categoría
                    df_final[col] = pd.to_numeric(df_final[col], errors='coerce').fillna(0.0)
            
            logger.info(f"DataFrame final creado: {len(df_final)} filas x {len(df_final.columns)} columnas")
            
            # Verificar si el archivo existe y está abierto, intentar eliminarlo
            if os.path.exists(ruta_completa):
                try:
                    os.remove(ruta_completa)
                    logger.info(f"Archivo existente eliminado: {ruta_completa}")
                except Exception as e:
                    logger.warning(f"No se pudo eliminar archivo existente: {e}")
                    # Agregar timestamp adicional si no se puede eliminar
                    import time
                    timestamp_extra = int(time.time())
                    nombre_archivo = f"Precios_Plu_{nombre_cadena.replace(' ', '_')}_{timestamp}_{timestamp_extra}.xlsx"
                    ruta_completa = os.path.join(ruta_salida, nombre_archivo)
                    logger.info(f"Usando nuevo nombre de archivo: {nombre_archivo}")
            
            # Crear archivo Excel con formato
            with pd.ExcelWriter(ruta_completa, engine='openpyxl') as writer:
                df_final.to_excel(writer, index=False, sheet_name='Precios')
                worksheet = writer.sheets['Precios']
                
                # Definir estilos
                header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                data_font = Font(name='Calibri', size=10)
                text_alignment = Alignment(horizontal='left', vertical='center')
                number_alignment = Alignment(horizontal='right', vertical='center')
                
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Aplicar formato a encabezados
                for col_num, column_title in enumerate(df_final.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                    
                    # Ajustar ancho de columnas
                    column_letter = get_column_letter(col_num)
                    max_length = 0
                    
                    # Encontrar longitud máxima en la columna
                    for cell_in_col in worksheet[column_letter]:
                        try:
                            if len(str(cell_in_col.value)) > max_length:
                                max_length = len(str(cell_in_col.value))
                        except:
                            pass
                    
                    # Establecer ancho basado en el contenido
                    adjusted_width = min(max_length + 2, 50)
                    
                    # Anchos especiales para columnas específicas
                    if column_title in ['#PLU_NUM_PLU', '#PLU']:
                        adjusted_width = max(12, adjusted_width)
                    elif column_title == 'PRODUCTO':
                        adjusted_width = max(40, adjusted_width)
                    else:
                        # Para columnas dinámicas (cat_descripcion|cat_id)
                        adjusted_width = max(15, adjusted_width)
                    
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Aplicar formato a datos
                total_filas = len(df_final)
                for row_num in range(2, total_filas + 2):
                    for col_num in range(1, len(df_final.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = thin_border
                        
                        # Determinar alineación basada en tipo de columna
                        column_title = df_final.columns[col_num - 1]
                        
                        if column_title in ['PRODUCTO']:
                            cell.alignment = text_alignment
                            cell.font = data_font
                        elif column_title in ['#PLU_NUM_PLU', '#PLU']:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.font = data_font
                        else:  # Columnas de precios
                            cell.alignment = number_alignment
                            cell.font = data_font
                            cell.number_format = '#,##0.00'
                
                # Aplicar filtros automáticos
                worksheet.auto_filter.ref = worksheet.dimensions
                
                # Congelar paneles (encabezados y primeras 3 columnas)
                worksheet.freeze_panes = 'D2'
                
                # Autoajustar altura de filas
                for row in worksheet.iter_rows(min_row=1, max_row=total_filas + 1):
                    worksheet.row_dimensions[row[0].row].height = 20
            
            logger.info(f"Archivo Excel generado exitosamente: {ruta_completa}")
            logger.info(f"Tamaño del archivo: {os.path.getsize(ruta_completa) / 1024:.2f} KB")
            
            return ruta_completa
            
        except Exception as e:
            logger.error(f"Error al generar archivo Excel: {e}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return None
    
    def proceso_completo(
        self,
        nombre_cadena: str,
        canal_param: str = 'Canal',
        canal_ids: Optional[List[str]] = None
    ) -> Optional[str]:
        logger.info("="*60)
        logger.info(f"INICIANDO PROCESO COMPLETO PARA: {nombre_cadena}")
        logger.info("="*60)
        
        try:
            # Validar cadena
            if not validar_cadena(nombre_cadena):
                logger.error(f"Cadena no válida: {nombre_cadena}")
                return None
            
            # Obtener ID de cadena
            cdn_id = obtener_cdn_id(nombre_cadena)
            logger.info(f"PASO 1: cdn_id obtenido: {cdn_id}")
            
            # Conectar a BD
            if not self.conectar():
                logger.error("No se pudo establecer conexión a la base de datos")
                return None
            
            # Obtener categorías
            logger.info("PASO 2: Consultando categorías...")
            categorias_df = self.obtener_categorias_por_cadena(cdn_id)
            if categorias_df is None or len(categorias_df) == 0:
                logger.error("PASO 2: No se pudieron obtener categorías")
                self.desconectar()
                return None
            logger.info(f"PASO 2: {len(categorias_df)} categorías obtenidas")
            
            # Ejecutar stored procedure
            logger.info("PASO 3: Ejecutando stored procedure...")
            df_precios = self.ejecutar_stored_procedure_precios(
                cdn_id, categorias_df, canal_param, canal_ids
            )
            if df_precios is None or df_precios.empty:
                logger.error("PASO 3: No se obtuvieron datos de precios")
                self.desconectar()
                return None
            logger.info(f"PASO 3: Datos obtenidos: {len(df_precios)} registros")
            
            # Generar archivo Excel
            logger.info("PASO 4: Generando archivo Excel...")
            ruta_archivo = self.generar_archivo_excel(df_precios, categorias_df, nombre_cadena, cdn_id)
            if ruta_archivo is None:
                logger.error("PASO 4: Error al generar archivo Excel")
                self.desconectar()
                return None
            
            logger.info(f"PASO 4: Archivo generado exitosamente")
            logger.info("="*60)
            logger.info("PROCESO COMPLETADO EXITOSAMENTE")
            logger.info(f"Archivo: {ruta_archivo}")
            logger.info("="*60)
            
            return ruta_archivo
            
        except Exception as e:
            logger.error(f"Error en proceso completo: {e}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return None
            
        finally:
            self.desconectar()

def procesar_cadena_simple(
    nombre_cadena: str,
    canal_param: str = 'Canal',
    canal_ids: Optional[List[str]] = None
) -> Optional[str]:
    consultas = ConsultasDB()
    return consultas.proceso_completo(nombre_cadena, canal_param, canal_ids)